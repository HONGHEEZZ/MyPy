
#------------------------------------------------------------------------------
#* 1. 파일 쓰기
#------------------------------------------------------------------------------
from openpyxl import Workbook
import time


wb = Workbook()

ws = wb.active

ws2 = wb.create_sheet("new_sheet2") #마지막 시트에 추가
ws1 = wb.create_sheet("new_sheet1", 1) #두번째 시트에 삽입

print(wb.sheetnames)

ws = wb['Sheet']
ws.title = '주소'

ws['A1'] = '이름'
ws['B1'] = '전화번호'

ws['A2'] = '한홍희'
ws['B2'] = '멋쟁이'

ws.cell(row = 3, column = 1, value='한가윤')
ws.cell(row = 3, column = 2, value='얘도 멋재이ㅌㅌ...')

ws.cell(4, 1).value='강영란'
ws.cell(4, 2).value='멋재이...'

cell = ws.cell(5, 1)
cell.value = '한지오'
wb.save('d:/hanhonghee.xlsx')


#------------------------------------------------------------------------------
#* 2. 파일 읽기
#------------------------------------------------------------------------------
from openpyxl import load_workbook
wb2 = load_workbook('d:/hanhonghee.xlsx')

print(wb2.sheetnames)

sheet = wb2['주소']
myval = sheet['A1'].value
print(myval)

for row in sheet.iter_rows():
    for cell in row:
        print(cell, cell.value)


#------------------------------------------------------------------------------
#* 3. 개별 셀 테이터 입력하기..
#------------------------------------------------------------------------------
from openpyxl import Workbook
import math
import datetime

wb = Workbook()
ws = wb.active

ws['A1'] = '한홍희' #문자열
ws['A2'] = 1234 #숫자 (int)
ws['A3'] = math.pi #숫자(float)
ws['A4'] = datetime.datetime(2031, 4, 16, 17, 52, 0) #시간 2021-04-16 17:52:00
ws['A5'] = '=SIN(PI()/2)' #수식
ws['A6'] = '=A1' #수식

wb.save('d:/openpyxl.xlsx')
