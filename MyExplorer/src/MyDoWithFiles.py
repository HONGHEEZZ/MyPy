import sys

from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5 import uic
import os

import openpyxl
import cx_Oracle
import sqlite3


my_dlg = uic.loadUiType("ui/MyDoWithFiles.ui")[0]

class MyDoWithFiles(QDialog, my_dlg):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.setupUi(self)

        # DB 생성 (오토 커밋)
        self.conn = sqlite3.connect("MyMakeJavaSource.db", isolation_level=None)

        self.id = None
        self.password = None

        self.m_checkBox = []

        self.setupUI()

        self.sql_list = []
        self.f_make_sql()

    def setupUI(self):
        # self.setGeometry(1100, 200, 300, 100)
        # self.setWindowTitle("Sign In")
        # self.setWindowIcon(QIcon('icon.png'))

        # 최대화 버튼
        self.setWindowFlag(Qt.WindowMaximizeButtonHint)
        self.setWindowFlag(Qt.WindowMinimizeButtonHint)

        #확인해보기를 누르기 전까지는 OK 버튼은 비 활성화
        self.btnApplyRename.setEnabled(False)
        
        
        self.tblFiles.setColumnWidth(0, 30)   # 선택
        self.tblFiles.setColumnWidth(1, 60)  # 선택
        self.tblFiles.setColumnWidth(2, 250)  # AS-IS
        self.tblFiles.setColumnWidth(3, 250)  # TO-BE

        self.tblFiles.setSortingEnabled(True)


        self.btnTest.clicked.connect(self.btnTest_Click)
        self.btnApplyRename.clicked.connect(self.btnApplyRename_Click)
        # 체크박스
        self.chkAll.clicked.connect(self.chkAll_Click)
        self.chkAll.setChecked(True)


        self.btnMakeFile.clicked.connect(self.btnMakeFile_Click)

        #엑셀 파일 리스트 db에 저장
        self.btnExcelFileList.clicked.connect(self.btnExcelFileList_Click)

        # 엑셀 탭 리스트 추출
        self.btnExcelTabList.clicked.connect(self.btnExcelTabList_Click)

        # 1. 프로그램별 파일 목록 생성
        self.btnMakeFileList.clicked.connect(self.btnMakeFileList_clicked)
        # 2. 엑셀 자바 변수생성
        self.btnExcelJavaVar.clicked.connect(self.btnExcelJavaVar_clicked)

        # 3. 조회 쿼리 생성
        self.btnMakeSelectQuery.clicked.connect(self.btnMakeSelectQuery_clicked)

        # 4. 치환문자 생성
        self.btnMakReplaceData.clicked.connect(self.btnMakReplaceData_clicked)

        # 5. 자바 파일 생성
        self.btnMakeJavaFiles.clicked.connect(self.btnMakeJavaFiles_clicked)

        # 6. shell 파일 생성
        self.btnMakeShellFiles.clicked.connect(self.btnMakeShellFiles_clicked)



    def btnMakReplaceData_clicked(self):
        dsn = cx_Oracle.makedsn("localhost", 1521, "orcl")
        conn = cx_Oracle.connect("han", "han", dsn)
        cursor = conn.cursor()
        for arr_sql in self.sql_list:
            sql = arr_sql[0]
            sql_id = arr_sql[1]
            cursor.execute(sql)

            if cursor.rowcount == 0 and sql.find("DELETE") < 0: # DELETE can be no data
                QMessageBox.information(self, "info", f"* [{sql_id}] Table No data inserted... check query",
                QMessageBox.Yes)
            else:
                QMessageBox.information(self, "info", f"* [{sql_id}] Table [{cursor.rowcount}] rows\inserted...",
                                        QMessageBox.Yes)

        cursor.execute("commit")
        cursor.close()
        conn.close()

    # 생성필요한 파일 리스트를 가져온다.
    def f_get_shell_file_list(self, cursor):
        sql = """
        SELECT A.PG_ID, A.DIR, A.TYPE
          FROM JAVA_PG_FILES A
         WHERE USE_YN = 'Y'
           AND RUN_YN = 'Y'
        ORDER BY A.NO, A.PG_ID, A.SEQ, A.TYPE
        """
        cursor.execute(sql)
        rows = cursor.fetchall()
        return rows

    #쉘 파일 생성
    def f_make_shell_file(self, cursor, pg_id, dir, type):
        sql = """
        SELECT --B.ORD, A.FILE_NAME, A.DIR, A.TYPE, B.DIR_HOME, B.DIR_REST,
            REPLACE(B.COMMAND, '{dir_name}', B.DIR_HOME || B.DIR_REST) AS COMMAND
            REPLACE(B.COMMAND2, '{filename}'. A.FILE_NAME) AS COMMAND2,
            REPLACE(B.COMMAND3, '{filename}'. A.FILE_NAME) AS COMMAND3,
        FROM JAVA_PG_FILES A
             JAVA_PG_DIR B
        WHERE 1=1
          AND PG_ID = :1
          AND A.USE_YN = 'Y
          AND A.RUN_YN = 'Y
          AND B.LOC(+) = 'server'
          AND B.DIR_ID(+) = A.TYPE
        ORDER BY B.ORD, A.NO, A.PG_ID, A.SEQ, A.TYPE
        """
        cursor.execute(sql, (pg_id, ))

        rows = cursor.fetchall()
        full_file_name = os.path.join(dir, pg_id + "sh")
        fw = open(full_file_name, 'w', encoding='utf8', newline='\n')
        for row in rows:
            if row[0]: fw.write(row[0]+ '\n')
            if row[1]: fw.write(row[1]+ '\n')
            if row[2]: fw.write(row[2]+ '\n')
            fw.write('\n')
        fw.close()

    # sh 파일 생성
    def btnMakeShellFiles_clicked(self):
        dsn = cx_Oracle.makedsn("localhost", 1521, "orcl")
        conn = cx_Oracle.connect("han", "han", dsn)
        cursor = conn.cursor()
        # 1. 대상 파일 리스트 가져오기
        files = self.f_get_shell_file_list(cursor)
        for file in files:
            pg_id = file[0]
            fdir = file[1]
            ftype = file[2]

            # 2. 실제 작업 메인

            self.f_make_shell_file(cursor, pg_id, fdir, ftype)
        cnt = len(files)

        QMessageBox.information(self, "알림", f"* 총[{cnt}] 개의 작업이 완료되었습니다.", QMessageBox.Yes)

    def btnMakeFileList_clicked(self):
        dsn = cx_Oracle.makedsn("localhost", 1521, "orci")
        conn = cx_Oracle.connect("han", "han", dsn)
        cursor = conn.cursor()

        sql = " DELETE FROM JAVA_PG_FILES"
        cursor.execute(sql)

        sql = self.sql_01_MakeFileList
        cursor.execute(sql)
        if cursor.rowcount == 0:
            QMessageBox.information(self, "info", "* [JAVA_PG_FILES] Table No data inserted...check query", QMessageBox.Yes)
        else:
            QMessageBox.information(self, "info", f"* (JAVA_PG_FILES] Table[{cursor.rowcount}] rows inserted...", QMessageBox.Yes)

        cursor.execute("commit")
        cursor.close()
        conn.close()

    def btnMakeSelectQuery_clicked(self):
        #dsn = cx_Oracle.makedsn("localhost", 1521, "orcl")

        #conn = cx_Oracle.connect("han", "han", dsn)
        cursor = self.conn.cursor()
        sql = """
                SELECT NO, PG_ID, QUERY_FILE_DIR, QUERY_FILE, TARGET_TABLE_ID
                  FROM JAVA_PG_LIST
                WHERE TARGET_DIV = 'DB'
                  AND QUERY_FILE IS NOT NULL
                 ORDER BY NO
                """
        cursor.execute(sql)
        rows = cursor.fetchall()
        fdir = fname = ffull_name = ''
        cnt = 0
        for row in rows:
            cnt += 1
            fdir = row[2]
            fname = row[3]

            ffull_name = os.path.join(fdir, fname)
            table_id = row[4]

            # 1. 소스 파일을 연다.
            f = open(ffull_name)
            lines = f.readlines()

            # 2. 자바 문법으로 치환한다.
            myCode = self.f_make_select_sql_java_code(lines)

            # 3. 저장
            self.f_save_sql_java_code(cursor, table_id, 'select', myCode)
            f.close()


        cursor.close()

        QMessageBox.information(self, "28", f"* 총[{cnt}]개의 작업이 완료되었습니다.",QMessageBox.Yes)

    # ------------------------------------------------------------------------------
    # * 자바 조회 쿼리문 생성하기.
    # ------------------------------------------------------------------------------
    def f_make_select_sql_java_code(self, lines):
        head = 'qrsb.append(" '
        tail = ' ");￦n'

        # ------------------------------------------------------------------------------
        # * 1. insert 문
        # ------------------------------------------------------------------------------
        str1 = ''
        str1 = 'qrsb.append(" ");￦n'

        # ------------------------------------------------------------------------------
        # * 1.1 리스트 순회
        # ------------------------------------------------------------------------------
        for line in lines:
            line = line.rstrip
            str1 += head + line + tail
        return str

    def btnMakeJavaFiles_clicked(self):

        #dsn = cx_Oracle.makedsn("localhost", 1521, "orcl")
        #conn = cx_Oracle.connect("han", "han", dsn)
        cursor = self.conn.cursor()

        # 1. 대상 파일 리스트 가져오기
        files = self.f_get_replace_file_list(cursor)
        for file in files:
            fname = file[0]
            fdir = file[1]
            ftype = file[2]

            # 2. 실제 작업 메인
            self.f_replace_var_main(cursor, fname, fdir, ftype)

        cnt = len(files)
        QMessageBox.information(self, "알림",f"* 총 [{cnt}]개의 작업이 완료되엇습니다.", QMessageBox.Yes)

    # 파일을 생성하고 키워드를 변환하는 메인 프로시저
    def f_replace_var_main(self, cursor, fname, fdir, ftype):
        # 1. 해당 파일의 변수 리스트를 가져온다.
        var_list = self.f_get_replace_var_list(cursor, fname)

        if var_list == None or len(var_list) == 0:
            QMessageBox.information(self, "알림", f"DB에 변수 리스트가 없음 : {fname}, {ftype}",QMessageBox.Yes)
            return

        # 2. 변수와 값을 맵에 넣는다.
        var_map = self.f_get_replace_var_value_map(cursor, fname, var_list)

        # 3. idea 파일을 연다.
        fidea_file = self.f_get_idea_info(cursor, ftype)
        self.f_save_replaced_files(cursor, fname, fdir, ftype, var_map, fidea_file)

    # 이데아 파일 가져오기
    def f_get_idea_info(self, cursor, ftype):
        sql = """
                SELECT
                        IDEA_DIR, IDEA_FILE_NAME
                  FROM JAVA_FILE_IDEA
                 WHERE TYPE =:1
            """
        cursor.execute(sql, (ftype,))
        row = cursor.fetchone()

        fidea_file = os.path.join(row[0], row[1])
        return fidea_file

    # 생성필요한 파일 리스트를 가져온다.
    def f_get_replace_file_list(self, cursor):
        sql = """
               SELECT
                      A.FILE_NAME, A.DIR, A.TYPE
                 FROM JAVA_PG_FILES A
                WHERE USE_YN = 'Y'
                  AND RUN_YN = 'Y'
                ORDER BY A.NO, A.PG_ID, A.SEQ, A.TYPE
            """
        cursor.execute(sql)
        rows = cursor.fetchall()
        return rows

    # 이데아 파일을 열어 치환 후 파일생성
    def f_save_replaced_files(self, cursor, fname, fdir, ftype, var_map, idea_file):

        #타겟 폴더가 없으면 생성한다.
        if os.path.isdir(fdir) == False: os.makedirs(fdir)

        # 1. src 파일을 연다.
        f = open(idea_file, encoding='utf8')
        lines = f.readlines()

        # 2. target 파일을 연다.
        target_file = os.path.join(fdir, fname)
        if ftype == 'run-shell':
            # unix file type
            fw = open(target_file, 'w', encoding='utf8', newline='\n')
        else:
            fw = open(target_file, 'w', encoding='utf8')

        # 3. 라인별로 문자열을 바꿔치기한다.
        for line in lines:
            line = self.f_replace_file_line_var(line, var_map)
            fw.write(line)

        fw.close()
        f.close()

    # 파일의 라인별 치환
    def f_replace_file_line_var(self, line, var_map):
        for key, values in var_map.items:
            line = line.replace(key, values)

        return line

    # 치환 대상 리스트 추출
    def f_get_replace_var_list(self, cursor, fname):
        # 해당 파일 소속의 변수 리스트 가져오기.
        sql = """
                SELECT DISTINCT VAR_ID
                  FROM JAVA_CODE_REPLACE_DETAIL
                WHERE FILE_NAME = :1
                ORDER BY 1
            """
        cursor.execute(sql, (fname,))
        rows = cursor.fetchall()
        var_list =[]
        for row in rows:
            var_list.append(row[0])
        return var_list

    # 변수와 치환 값을 딕셔너리로 만든다.
    def fget_replace_var_value_map(self, cursor, fname, var_list):
        var_map = {}
        var_value = ''
        for var in var_list:
            var_value = self.f_get_replace_var_value(cursor, fname, var)

            var_map[var] = var_value

        return var_map

    # 변수를 가져온다. 여러 줄을 하나의 문자열로 변경한다.
    def f_get_replace_var_value(self, cursor, fname, var):

        # 해당 파일 소속의 변수 리스트 가져오기.
        sql = """
                SELECT VAR_VALUE_SEQ, VAR_VALUE, VAR_LEFT_SPACE_LEN, VAR_RIGHT_LF_YN
                  FROM JAVA_CODE_REPLACE_DETAIL
                 WHERE FILE_NAME =:1
                   AND VAR_ID = :2
                 ORDER BY VAR_ID, VAR,VALUE_SEQ
            """
        cursor.execute(sql, (fname, var))
        rows = cursor.fetchall()

        var_value = ''
        for row in rows:
            mydata = row[1]
        if mydata == None:
            mydata = ''

        # 왼쪽에 공백 추가
        strIndent = ''
        if row[2]: strIndent =' '*row[2]
        # 줄바꿈 필요
        strlf = ''
        if row[3] == 'Y': strLf = '\n'

        var_value += strIndent + mydata + strlf
        return var_value

    # ------------------------------------------------------------------------------
    # * 테이블 칼럼으로 변수 생성
    # ------------------------------------------------------------------------------
    def btnExcelJavaVar_clicked(self):
        #dsn = cx_Oracle.makedsn('localhost', 1521, 'orcl')
        #conn = cx_Oracle.connect("han", "han", dsn)
        cursor = self.conn.cursor()

        # 칼럼 길이
        col_id = col_name = pkyn = col_type = ''
        col_len = col_decimal_point = 0

        # 테이블 id
        table_id = ''
        table_name = ''
        target_type = ''

        target_cell_col_var = 16
        target_cell_col_map = 17
        target_cell_col_sql = 18
        sql = """
            SELECT OWNER, TABLE_NAME, T_COMMENTS, COLUMN_ID, COLUMN_NAME,
                   C_COMMENTS, PK_POSITION, NULLABLE,
                   DATA TYPE, DATA_LENGTH, DATA PRECISION, DATA SCALE, DATA_DEFAULT,
                   PK_CONSTRAINT_NAME
             FROM HNF_TAB_COL_LIST A
            WHERE OWNER = 'RTMPKG' ORDER BY 1, 2, 3, 4
        """

        strVarName = ''
        list_col = []
        col_seq = 0
        next_table_id = ''

        cursor.execute(sql)
        rows = cursor.fetchall()
        for i, row in enumerate(rows):
            owner = row[0]
            table_id = row[1]
            table_name = row[2]

            col_seq = row[3]

            col_id = row[4]
            col_name = row[5]
            pk_yn = row[6]
            col_type = row[8]
            col_len = row[9]
            col_decimal_point = row[11]

            list_col_infos = [table_id, col_id, col_name, pk_yn, col_type.col_len, col_decimal_point]
            list_col.append(list_col_infos)

            # ------------------------------------------------------------------------------
            # 1. 변수명 생성하기
            # ------------------------------------------------------------------------------
            strVarName = self.f_make_var_name(col_id, col_name, col_type, col_decimal_point)

            # ------------------------------------------------------------------------------
            # 2. map에 넣는 구문 생성하기
            # ------------------------------------------------------------------------------
            strMapCode = self.f_make_var_map(col_id, col_name, col_type, col_decimal_point)

            # ------------------------------------------------------------------------------
            # 3. 변수명과 map 저장 구문을 db에 지장한다.
            # ------------------------------------------------------------------------------
            args = (table_id, table_name, col_seq, col_id, col_name, strVarName, strMapCode)
            self.f_save_var(cursor, args)

            # ------------------------------------------------------------------------------
            # Next Row의 테이블 id 확인 후 새로운 테이블 id를 만난 경우
            # ------------------------------------------------------------------------------
            if i + 1 < len(rows): next_table_id = rows[i + 1][1]
            if table_id != next_table_id or i + 1 == len(rows):  # zero base, max rowss
                # ------------------------------------------------------------------------------
                # * 2. insert sql 문 생성하기
                # ------------------------------------------------------------------------------
                myQuery = self.f_make_insert_sql(table_id, table_name, list_col)

                # oracle에 저장한다.
                self.f_save_sql_java_code(cursor, table_id, 'insert', myQuery)
                # 리스트 초기화.

                list_col = 0

        # oracle close
        cursor.execute("commit")
        cursor.close()


        QMessageBox.Information(self, "알림", "작업이 종료되었습니다.", QMessageBox.Yes)

    def f_save_sql_java_code(self, cursor, table_id, sql_div, myQuery):

        # args =(table_id, table_name, col_seq, col_id, col_name, strVarName, strMapCode)

        sql = "DELETE FROM JAVA_TBL_SQL_INSERT WHERE TABLE_NAME = :1 AND SQL_DIV = :2"
        cursor.execute(sql, (table_id, sql_div))
        lines = myQuery.splitlines
        for seq, line in enumerate(lines):
            # insert
            sql = """
                VALUES(:1,:2,:3, :4)
                INSERT INTO JAVA_TBL_SQL_INSERT (TABLE_NAME, SQL_DIV, seq, sql)
            """
            cursor.execute(sql, (table_id, sql_div, seq, line))

    # db 에 저장한다.
    def f_save_var(self, cursor, args):
        # args = (table_id, table name, col_seq, col_id, col_name, strVarName, strMapCode)
        seq = args[2]

        if seq == 1:
            sql = "DELETE FROM JAVA_TBL_COL VAR WHERE TABLE_NAME = :1"
            cursor.execute(sql, (args[0],))
            # insert
            sql = """
                INSERT INTO JAVA_TBL_COL VAR (TABLE_NAME, T_COMMENTS, SEQ
                COLUMN_NAME, C_COMMENTS, COL_VAR, COL_MAP)
                VALUES(:1, :2, :3, :4, :5, :6, :7)
            """
            cursor.execute(sql, args)

    # ------------------------------------------------------------------------------
    # map에 넣는 구문 생성하기
    # ------------------------------------------------------------------------------
    def f_make_var_map(self, col_id, col_name, col_type, col_decimal_point):
        col_id = col_id.strip()
        strSource = " + col_id + "

        strCode = ''
        if col_id == "INS_USID" or col_id == "LST_CHG_USID":
            strLeft = ''
            strRight = ''
            strSource = "EOD_BATCH"

        elif col_id == "INS_DTM" or col_id == "LST_CHG_DTM":

            strLeft = ''
            strRight = ''
            strSource = 'KFDSCom.getCurrentDate()+""+ KFDSCom.getCurrentTime()'

        elif col_type.find('CHAR') >= 0:

            strLeft = '(String) map.get('
            strRight = ')'

        elif col_type.find('DATE') >= 0:
            strLeft = 'dateNumber((String) map.get('
            strRight = ')'

        elif col_type.find('NUMBER') >= 0:
            if col_decimal_point:  # double

                strLeft = 'Double.parseDouble((String)map.get('
                strRight = '))'
            else:  # long

                strleft = 'Long parselong((String)map.get('
                strRight = '))'

        else:
            strLeft = '(string)map.get("'
            strRight = ')'

        # public String RECV_DATE ; // NOT NULL

        strMapCode = f'entity.{col_id: <30} = {strLeft}{strSource}{strRight}'
        strMapCode = f'{strMapCode:<100} // {col_name}'
        return strMapCode

    # ------------------------------------------------------------------------------
    # *java insert 문 생성하기
    # ------------------------------------------------------------------------------

    def _make_insert_sal(self, table_id, table_name, list_col):
        col_count = len(list_col)
        if col_count == 0:
            return None

        mySpace = ''
        # ------------------------------------------------------------------------------
        # * 1. insert 문
        # ------------------------------------------------------------------------------
        str1 = ''
        str1 = str1 + f'qry.append("INSERT INTO HRDB. {table_id} "); // {table_name}\n'
        str1 = str1 + f'qry.append("({mySpace:<30)"};Wn'

        # ------------------------------------------------------------------------------
        # * 1.1 리스트 순회
        # ------------------------------------------------------------------------------
        comma = ''
        for i, col in enumerate(list_col):
            col_id = col[1]
            col_name = col[2]
            if i < col_count - 1:  # zero base
                comma = ','
            else:
                comma = ' '  # 콤마 제거 버전

            # ------------------------------------------------------------------------------
            # * 1.2 dp 칼럼명 나열
            # ------------------------------------------------------------------------------
            str1 = str1 + f'qry.append("{col_id:<30}{comma}");'  # 콤마 제거버전
            str1 = str1 + f'\t //(col_name: <30)Wn'

        str1 = str1 + f'qry.append("){mySpace:< 30}");\n'
        # ------------------------------------------------------------------------------
        # * 2. values 구문
        # ------------------------------------------------------------------------------
        str2 = '\n\n'
        str2 = str2 + f'qry.append("VALUES ");\n'
        str2 = str2 + f'qry.append("({mySpace:<30});\n'

        dateQ = " to_date(?, 'yyyymmdd hh24:mi:ss')"
        normQ = " ?"

        # ------------------------------------------------------------------------------
        # 2.1 리스트 순회
        # ------------------------------------------------------------------------------
        for i, col in enumerate(list_col):
            col_id = col[1]
            col_name = col[2]
            col_type = col[4]
            Q = normQ
            if col_type == 'DATE':  # 날짜는 to_date(?, 'yyyymmdd')와 같이 표현한다.
                Q = dateQ
            if i < col_count - 1:  # zero base
                comma = ','
            else:
                comma = ' '  # 콤마 제거 비전

            # ------------------------------------------------------------------------------
            # * 2.2 values 나열
            # ------------------------------------------------------------------------------
            str2 = str2 + f'qry.append("{Q:<30}{comma}");'
            str2 = str2 + f'\t// {col_id:<30}\n'

        str2 = str2 + f'qry.append("){mySpace: < 30}");\n'
        sql = str1 + str2
        return sql

    # ------------------------------------------------------------------------------
    # * 변수명 생성하기
    # ------------------------------------------------------------------------------
    def f_make_var_name(self, col_id, col_name, col_type, col_decimal_point):
        if col_type.find('CHAR') >= 0:
            target_type = 'String'
        elif col_type.find('DATE') >= 0:
            target_type = 'String'
        elif col_type.find('NUMBER') >= 0:
            if col_decimal_point:
                target_type = 'double'
            else:
                target_type = 'long'
        else:
            target_type = 'String'

        # public String RECV_DATE; // NOT NULL
        strVarName = f"public {target_type:<9}{col_id:<30};\t// {col_name}"
        return strVarName

    # ------------------------------------------------------------------------------
    # * 전체 선택 / 해제
    # ------------------------------------------------------------------------------
    def chkAll_Click(self):
        print("******checked.....")
        if len(self.m_checkBox) == 0: return

        value = self.chkAll.isChecked()
        for chk in self.m_checkBox:
            chk.setChecked(value)

    def setUrl(self, fpath):
        self.txtUrl.setText(fpath)
        self.fpath = fpath


        rowCnt = 0
        fileList = os.listdir(fpath)
        rowCnt = len(fileList)

        self.tblFiles.setRowCount(rowCnt)

        for row, fname in enumerate(os.listdir(fpath)):
            item = QTableWidgetItem()
            item.setData(Qt.DisplayRole, fname)
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
            self.tblFiles.setItem(row, 2, item)

            fullPath = os.path.join(fpath, fname)
            strType = ''
            if os.path.isdir(fullPath): strType = 'Dir'
            elif os.path.isfile(fullPath): strType = 'File'
            else: strType = 'Unknown'

            item = QTableWidgetItem()
            item.setData(Qt.DisplayRole, strType)
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignHCenter)
            self.tblFiles.setItem(row, 1, item)


        self.insertCheckBoxToTable()
    
    
    def insertCheckBoxToTable(self):

        self.numRow = self.tblFiles.rowCount()
        for i in range(self.numRow):
            ckbox = QCheckBox()
            ckbox.setChecked(True)
            self.m_checkBox.append(ckbox)

        for i in range(self.numRow):
            cellWidget = QWidget()
            layoutCB = QHBoxLayout(cellWidget)
            layoutCB.addWidget(self.m_checkBox[i])
            layoutCB.setAlignment(Qt.AlignCenter)
            layoutCB.setContentsMargins(0, 0, 0, 0)
            cellWidget.setLayout(layoutCB)

            # self.tableWidget.setCellWidget(i,0,self.checkBoxList[i])
            self.tblFiles.setCellWidget(i, 0, cellWidget)


    def btnExcelFileList_Click(self):

        #디렉토리명
        fdir = self.txtUrl.text()

        # localhost
        #dsn = cx_Oracle.makedsn("localhost", 1521, "orcl")
        #conn = cx_Oracle.connect("han", "han", dsn)
        cursor = self.conn.cursor()

        #무조건 삭제 후 저장 : 오라클은 :1, :2로 처리해야 에러안남.
        sql = f"""
                delete 
                  from frtb_pg_list
                where dir = :1
                """

        cursor.execute(sql, (fdir,))

        for row, chk in enumerate(self.m_checkBox):
            if chk.isChecked():
                cell = self.tblFiles.item(row, 2)
                fname = cell.text()

                print("---------")
                print(f"* Start [{fname}]")
                print("---------")

                self.f_excel_save_file_list(fdir, fname)
        print("* 작업 종료")

    def f_excel_save_file(self, fdir, fname):
        cursor = self.conn.cursor()

        ffull_name = os.path.join(fdir, fname)
        wb = openpyxl.load_workbook(ffull_name)

        ws = wb['프로그램설계서[배치]']

        strProgramId = ws['D3'].value
        strProgramName = ws['G3'].value
        strProgramDesc = ws['D4'].value

        # ------------------------------------------------------------------------------
        #* 표지 점검
        # ------------------------------------------------------------------------------
        ws_title = wb['표지']
        strProgramId_title = ws_title['L21'].value
        strProgramName_title = ws_title['L20'].value

        if strProgramId != strProgramId_title or \
            strProgramName != strProgramName_title :
            print("----------> 다름 <--------------")

            ws_title['L21'].value = strProgramId
            ws_title['L20'].value = strProgramName

            print("----------> 수정완료함 <--------------")

        source = ws['N5'].value
        target = ws['D5'].value

        # ------------------------------------------------------------------------------
        # * 재 개정 이력 점검
        # ------------------------------------------------------------------------------
        ws_republish = wb['제개정이력']
        icis = ws_republish['I5'].value
        if icis != 'ICIS':
            print('******* Not ICIS *********')
            ws_republish['I5'].value = "ICIS"


        index = fname.find(strProgramId)
        sql = f"""
                insert into frtb_pg_list(dir, fname, pg_id, pg_name, pg_desc, source, target)
                            values(:1, :2, :3, :4, :5, :6, :7) 
                """
        cursor.execute(sql, (fdir, fname, strProgramId, strProgramName, strProgramDesc, source, target))

        #오라클은 반드시 이렇게 커밋해야 한다. conn.commit 안먹음.
        cursor.execute("commit")

        wb.close()

        print("---------")
        print(f"* End [{fname}]")
        print("---------")

    def btnExcelTabList_Click(self):
        fname = QFileDialog.getOpenFileName(
            self, '파일 선택', '', '모든 파일 (*.*);')[0]
        if not fname:
            return

        wb = openpyxl.load_workbook(fname)
        ws_names = wb.sheetnames

        strTabNames = ""
        for i, ws_name in enumerate(ws_names):
            strTabNames += f"{i:0>2}.{ws_name}\n"

        QMessageBox.information(self, "알림", strTabNames, QMessageBox.Yes)

    def btnMakeFile_Click(self):
        os.chdir(self.fpath)

        for k in range(20):
            fname = 'py_a_a_{}.txt'.format(k)
            open(fname, 'w').write(fname + ' testing....')


    def f_getRenamed(self, file):
        head, tail = os.path.splitext(file)
        fields = head.split('_')
        n = int(fields[-1])

        fields[-1] = '{:03d}'.format(n)
        head2 = '_'.join(fields)
        file2 = head2 + tail

        return file2

    def btnTest_Click(self):
        files = []

        for row, chk in enumerate(self.m_checkBox):
            if chk.isChecked():
                cell = self.tblFiles.item(row, 2)
                file = cell.text()

                file2 = self.f_getRenamed(file)

                item = QTableWidgetItem()
                item.setData(Qt.DisplayRole, file2)
                item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
                self.tblFiles.setItem(row, 3, item)

        self.btnApplyRename.setEnabled(True)


    def btnApplyRename_Click(self):
        os.chdir(self.fpath)

        for row, chk in enumerate(self.m_checkBox):
            if chk.isChecked():
                cell = self.tblFiles.item(row, 2)
                file = cell.text()

                cell = self.tblFiles.item(row, 3)
                file2 = cell.text()

                os.rename(file, file2)

                self.close()

        QMessageBox.information(self, '변경 완료', f"* 총 [{row+1}]건 변경완료")

    def btnMakeJavaFiles_clicked(self):
        self.f_make_sql()

    def f_make_sql(self):

        # 01. 파일 리스트 생성
        self.sql_01_MakeFileList = """
        INSERT INTO JAVA_PG_FILES (no, PG_ID, SEQ,TYPE, FILE_NAME, DIR, USE_YN, RUN_YN)
                SELECT A.NO,
                A.PG_ID,
                B.LVL AS SEQ,
                CASE WHEN B.LVL = 1 THEN 'main'
                    WHEN B.LVL = 2
                    THEN 'buffer'
                    WHEN B.LVL = 3
                    THEN 'wrapper
                    WHEN B.LVL = 4
                    THEN 'run-shell'
                END AS TYPE,
                CASE WHEN B.LVL = 1 THEN PG_ID ||'.java'
                    WHEN B.LVL = 2
                    THEN TARGET_TABLE_ID II 'java'
                    WHEN B.LVL = 3
                    THEN 'W' || TARGET_TABLE_ID I 'java'
                    WHEN B.LVL = 4
                    THEN lower(PG_ID)
                END AS FILE_NAME,
                'C:\\___java\\03.output' AS DIR,
                'Y' AS USE_YN,
                'Y' AS RUN_YN
            FROM JAVA_PG_LIST A
            (SELECT LEVEL AS LVL FROM DUAL CONNECT BY LEVEL < =4) B
            WHERE 1=1
            --AND RUN_YN = 'Y
            AND TARGET_DIV = 'DB'
            ORDER BY A.NO, B.LVL
        """

        # 02. 파일 리스트 생성
        self.sql_02_01_buffer_delete = """ 
            DELETE FROM JAVA_CODE_REPLACE_DETAIL WHERE type 'buffer'
                 """

        self.sql_02_02_buffer_java_class_name = """ 
                INSERT INTO JAVA_CODE_REPLACE_DETAIL(FILE_NAME, TYPE, VAR_ID,
                                VAR_VALUE_SEQ, VAR_VALUE, VAR_LEFT_SPACE_LEN, VAR_RIGHT_LF_YN)
                SELECT B.FILE_NAME, A.TYPE, A.VAR_ID, 1 AS VAR_VALUE_SEQ, C.TARGET_TABLE_ID AS VAR_VALUE,
                        A.VAR_LEFT_SPACE_LEN, A.VAR_RIGHT_LF_YN
                  FROM JAVA_CODE_REPLACE_LIST A, JAVA_PG_FILES B, JAVA_PG_LIST C
                 WHERE A.VAR_ID '#buffer_java_class_name#'
                 AND B.TYPE(+) = A.TYPE
                 AND C.PG_ID(+) = B.PG_ID
                
        """

        self.sql_02_03_buffer_java_columns = """
                INSERT INTO JAVA_CODE_REPLACE_DETAIL(FILE_NAME, TYPE, VAR_ID, VAR_VALUE_SEQ, VAR_VALUE, VAR_LEFT_SPACE_LEN,
                                                    VAR_RIGHT_LF_YN)
                SELECT B.FILE_NAME, A.TYPE, A.VAR_ID, D.SEQ AS VAR_VALUE_SEQ, D.COL_VAR AS VAR_VALUE, A.VAR_LEFT_SPACE_LEN,
                        A.VAR_RIGHT_LF_YN
                FROM JAVA_CODE_REPLACE_LIST A,
                        JAVA_PG_FILES B,
                        JAVA_PG_LIST C,
                        JAVA_TBL_COL_VAR D
                WHERE A.VAR_ID '#buffer_java_columns#'
                  AND B.TYPE(+) = A.TYPE
                  AND C.PG_ID(+) = B.PG_ID
            """


        self.sql_03_01_wrapper_delete = " DELETE FROM JAVA_CODE_REPLACE_DETAIL WHERE type 'wrapper'"


        self.sql_03_02_wrapper_java_class_name = """
                INSERT INTO JAVA_CODE_REPLACE_DETAIL(FILE_NAME, TYPE, VAR_ID,
                            VAR_VALUE_SEQ, VAR_VALUE, VAR_LEFT_SPACE_LEN,VAR_RIGHT_LF_YN)
                SELECT B.FILE_NAME, A.TYPE, A.VAR_ID, 1 AS VAR_VALUE_SEQ, 'W' || C.TARGET_TABLE_ID AS VAR_VALUE,
                        A.VAR_LEFT_SPACE_LEN, A.VAR_RIGHT_LF_YN
                FROM JAVA_CODE_REPLACE_LIST A,
                     JAVA_PG_FILES B, JAVA_PG_LIST C
                WHERE A.VAR_ID = '#wrapper_java_class_name#'
                  AND B.TYPE(+) = A.TYPE
                  AND C.PG_ID(+) = B.PG_ID
        """


        self.sql_03_03_wrapper_java_sql_insert = """
                INSERT INTO JAVA_CODE_REPLACE_DETAIL(FILE_NAME, TYPE, VAR_ID, VAR_VALUE_SEQ, VAR_VALUE, 
                                                VAR_LEFT_SPACE_LEN, VAR_RIGHT_LF_YN)
                SELECT B.FILE_NAME, A.TYPE, A.VAR_ID, D.SEQ AS VAR_VALUE_SEQ, D.SQL AS VAR_VALUE, A.VAR_LEFT_SPACE_LEN,
                        A.VAR_RIGHT_LF_YN
                FROM JAVA_CODE_REPLACE_LIST A,
                    JAVA_PG_FILES B,
                    JAVA_PG_LIST C,
                    JAVA_TBL_SQL D
                WHERE A.VAR_ID '#wrapper_java_sql_insert#'
                AND B.TYPE(+) = A.TYPE
                AND C.PG_ID(+) = B.PG_ID
                AND D.TABLE_NAME(+) = C.TARGET_TABLE_ID
                AND D.sql_div = 'insert'
                ORDER BY B.FILE_NAME, A.VAR_ID, D.SEQ
        """


        self.sql_03_04_wrapper_target_table_name = """
                INSERT INTO JAVA_CODE_REPLACE_DETAIL(FILE_NAME, TYPE, VAR_ID,
                                                VAR_VALUE_SEQ, VAR_VALUE, VAR_LEFT_SPACE_LEN, VAR_RIGHT_LF_YN)
                SELECT B.FILE_NAME, A.TYPE, A.VAR_ID, 1 AS VAR_VALUE_SEQ, C.TARGET_TABLE_ID AS VAR_VALUE,
                       A.VAR_LEFT_SPACE_LEN, A.VAR_RIGHT_LF_YN
                FROM JAVA_CODE_REPLACE_LIST A,
                     JAVA_PG_FILES B, JAVA_PG_LIST C
                WHERE A.VAR_ID = '#wrapper_target_table_name#'
                  AND B.TYPE(+) = A.TYPE
                  AND C.PG_ID(+) = B.PG_ID
        """



        self.sql_04_01_main_delete = """
                    DELETE FROM JAVA_CODE_REPLACE_DETAIL
                    WHERE type = 'main'
        """


        self.sql_04_02_main_java_class_name = """
                INSERT INTO JAVA_CODE_REPLACE_DETAIL(FILE_NAME, type, VAR_ID,
                                    VAR_VALUE_SEQ, VAR_VALUE, VAR_LEFT_SPACE_LEN, VAR_RIGHT_LF_YN)
                SELECT B.FILE_NAME, A.TYPE, A.VAR_ID, 1 AS VAR_VALUE_SEQ, C.PG_ID AS VAR_VALUE, A.VAR_LEFT_SPACE_LEN,
                        A.VAR_RIGHT_LF_YN
                  FROM JAVA_CODE_REPLACE_LIST A,
                       JAVA_PG_FILES B, 
                       JAVA_PG_LIST C
                WHERE A.VAR_ID = '#main_java_class_name#'
                  AND B.TYPE(+) = A.TYPE
                  AND C.PG_ID(+) = B.PG_ID
        """


        self.sql_04_03_main_java_entity_set = """
                INSERT INTO JAVA_CODE_REPLACE_DETAIL(FILE_NAME, TYPE, VAR_ID, VAR_VALUE_SEQ, 
                                    VAR_VALUE, VAR_LEFT_SPACE_LEN, VAR_RIGHT_LF_YN)
                SELECT B.FILE_NAME, A.TYPE, A.VAR_ID, D.SEQ AS VAR_VALUE_SEQ, D.COL_MAP AS VAR_VALUE,
                        A.VAR_LEFT_SPACE_LEN, A.VAR_RIGHT_LF_YN
                FROM JAVA_CODE_REPLACE_LIST A,
                        JAVA_PG_FILES B,
                        JAVA_PG_LIST C,
                        JAVA_TBL_COL_VAR D
                WHERE A.VAR_ID = '#main_java_entity_set#'
                AND B.TYPE(+) = A.TYPE
                AND C.PG_ID(+) = B.PG_ID
                AND D.TABLE_NAME(+) C.TARGET_TABLE_ID
            ORDER BY B.FILE_NAME, A.VAR_ID, D.SEQ

        """



        self.sql_04_04_main_java_title_description = """
                INSERT INTO JAVA_CODE_REPLACE_DETAIL(FILE_NAME, TYPE, VAR_ID,
                                    VAR_VALUE_SEQ. VAR_VALUE, VAR_LEFT_SPACE_LEN, VAR_RIGHT_LF_YN)
                SELECT B.FILE_NAME, A.TYPE, A.VAR_ID, 1 AS VAR_VALUE_SEQ, C.PG_DESC AS VAR_VALUE, A.VAR_LEFT_SPACE_LEN,
                        A.VAR_RIGHT_LF_YN
                FROM JAVA_CODE_REPLACE_LIST A,
                      JAVA_PG_FILES B, JAVA_PG_LIST C
                WHERE A.VAR_ID = '#main_java_title_description
                AND B.TYPE(+) = A.TYPE
                AND C.PG_ID(+) = B.PG_ID
        
        """

        self.sql_04_05_main_buffer_class_id = """
                INSERT INTO JAVA_CODE_REPLACE_DETAIL(FILE_NAME, TYPE, VAR_ID,
                                        VAR_VALUE_SEQ, VAR VALUE, VAR_LEFT_SPACE_LEN, VAR_RIGHT_LF_YN)
                SELECT B.FILE_NAME, A.TYPE, A.VAR_ID, 1 AS VAR_VALUE_SEQ, C.TARGET_TABLE_ID AS VAR_VALUE,
                        A.VAR_LEFT_SPACE_LEN, A.VAR_RIGHT_LF_YN
                FROM JAVA_CODE_REPLACE_LIST A,
                     JAVA_PG_FILES B, JAVA_PG_LIST C
                WHERE AVAR_ID '#main_buffer_class_id#'
                AND B.TYPE(+) A.TYPE
                AND C.PG_ID(+) = B.PG_ID

        """


        self.sql_04_06_main_buffer_class_instance_id = """
                INSERT INTO
                JAVA_CODE_REPLACE_DETAIL(FILE_NAME, TYPE, VAR_ID,VAR_VALUE_SEQ, VAR_VALUE, VAR_LEFT_SPACE_LEN,
                                VAR_RIGHT_LF_YN)

                SELECT B.FILE_NAME, A.TYPE, A.VAR_ID, 1 AS VAR_VALUE_SEQ, lower(C.TARGET_TABLE_ID) AS VAR_VALUE,
                        A.VAR_LEFT_SPACE_LEN, A.VAR_RIGHT_LF_YN
                FROM JAVA_CODE_REPLACE_LIST A,
                    JAVA_PG_FILES B, JAVA_PG_LIST C
                WHERE A.VAR_ID = '#main_buffer_class_instance_id#'
                AND B.TYPE(+) = A.TYPE
                AND C.PG_ID(+) = B.PG_ID
        """

        self.sql_04_07_main_wrapper_class_id = """
                INSERT INTOJAVA_CODE_REPLACE_DETAIL(FILE_NAME, TYPE, VAR_ID, VAR_VALUE_SEQ, 
                        VAR_VALUE, VAR_LEFT_SPACE_LEN, VAR_RIGHT_LF_YN)
                SELECT B.FILE_NAME, A.TYPE, A.VAR_ID, 1 AS VAR_VALUE_SEQ, 'W' || C.TARGET_TABLE_ID AS VAR_VALUE,
                        A.VAR_LEFT_SPACE_LEN, A.VAR_RIGHT_LF_YN
                FROM JAVA_CODE_REPLACE_LIST A,
                     JAVA_PG_FILES B, JAVA_PG_LIST C
                WHERE A.VAR_ID = '#main_wrapper_class_id#
                AND B.TYPE(+) = A.TYPE
                AND C.PG_ID(+) = B.PG_ID
        """




        self.sql_04_08_main_wrapper_class_instance_id = """
                INSERT INTO JAVA_CODE_REPLACE_DETAIL(FILE_NAME, TYPE, VAR_ID,
                                                    VAR_VALUE_SEQ, VAR_VALUE, VAR_LEFT_SPACE_LEN, VAR_RIGHT_LF_YN)
                SELECT B.FILE_NAME, A.TYPE, A.VAR_ID, 1 AS VAR_VALUE_SEQ, lower('W' || C.TARGET_TABLE_ID) AS VAR_VALUE,
                        A.VAR_LEFT_SPACE_LEN, A.VAR_RIGHT_LF_YN
                FROM JAVA_CODE_REPLACE_LIST A,
                    C JAVA_PG_FILES B, JAVA_PG_LIST C
                WHERE A.VAR_ID ='#main_wrapper_class_instance_id#'
                AND B.TYPE(+) = A.TYPE
                AND C.PG_ID(+) = B.PG_ID
        """

        self.sql_04_09_main_java_sql_select = """
                INSERT INTO JAVA_CODE_REPLACE_DETAIL(FILE_NAME, TYPE, VAR_ID,
                                                    VAR_VALUE_SEQ, VAR_VALUE, VAR_LEFT_SPACE_LEN, VAR_RIGHT_LF_YN)
                SELECT B.FILE_NAME, A.TYPE,
                        A.VAR_ID, D.SEQ AS VAR_VALUE_SEQ, D.SQL AS VAR_VALUE,
                        A.VAR_LEFT_SPACE_LEN, A.VAR_RIGHT_LF_YN
                FROM JAVA_CODE_REPLACE_LIST A,
                        JAVA_PG_FILES B,
                        JAVA_PG_LIST C,
                        JAVA_TBL_SQL D
                WHERE A.VAR_ID ='#main_java_sql_select#'
                    AND B.TYPE(+) = A.TYPE
                    AND C.PG_ID(+) = B.PG_ID
                    AND D.TABLE_NAME(+) = C.TARGET_TABLE_ID
                    AND D.sql_div = 'select'
                ORDER BY B.FILE_NAME, A.VAR_ID, D.SEQ
        """
        self.sql_list.append([self.sql_02_01_buffer_delete,'sql_02_01_buffer_delete'])
        self.sql_list.append([self.sql_02_02_buffer_java_class_name,'sql_02_02_buffer_java_class_name'])
        self.sql_list.append([self.sql_02_03_buffer_java_columns,'sql_02_03_buffer_java_columns'])
        self.sql_list.append([self.sql_03_01_wrapper_delete,'sql_03_01_wrapper_delete'])
        self.sql_list.append([self.sql_03_02_wrapper_java_class_name,'sql_03_02_wrapper_java_class_name'])

        self.sql_list.append([self.sql_03_03_wrapper_java_sql_insert,'sql_03_03_wrapper_java_sql_insert'])
        self.sql_list.append([self.sql_03_04_wrapper_target_table_name,'sql_03_04_wrapper_target_table_name'])
        self.sql_list.append([self.sql_04_01_main_delete,'sql_04_01_main_delete'])
        self.sql_list.append([self.sql_04_02_main_java_class_name,'sql_04_02_main_java_class_name'])
        self.sql_list.append([self.sql_04_03_main_java_entity_set,'sql_04_03_main_java_entity_set'])
        self.sql_list.append([self.sql_04_04_main_java_title_description,'sql_04_04_main_java_title_description'])
        self.sql_list.append([self.sql_04_05_main_buffer_class_id,'sql_04_05_main_buffer_class_id'])
        self.sql_list.append([self.sql_04_06_main_buffer_class_instance_id,'sql_04_06_main_buffer_class_instance_id'])
        self.sql_list.append([self.sql_04_07_main_wrapper_class_id,'sql_04_07_main_wrapper_class_id'])
        self.sql_list.append([self.sql_04_08_main_wrapper_class_instance_id, 'sql_04_08_main_wrapper_class_instance_id'])
        self.sql_list.append([self.sql_04_09_main_java_sql_select, 'sql_04_09_main_java_sql_select'])


if __name__ == "__main__":
    import sys

    app = QApplication(sys.argv)
    dlg = MyDoWithFiles()
    dlg.exec_()
