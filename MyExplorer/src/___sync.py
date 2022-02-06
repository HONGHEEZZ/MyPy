import sys

from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5 import uic
import os

import openpyxl
import cx_Oracle


# ------------------------------------------------------------------------------
# * 테이블 칼럼으로 변수 생성
# ------------------------------------------------------------------------------
def btnExcellavaVar_clicked(self):
    dsn = cx_Oracle.makedsn('localhost', 1521, 'orcl')
    conn = cx_Oracle.connect("han", "han", dsn)
    cursor = conn.cursor()

    # 칼럼 길이
    col_id = col_name = pkyn = col_type = ''
    col_len = col_decimal_point = 0

    # 테이블 id
    table_id = ''
    table_name = ''
    target_type = ''

    target_cell_col_var = 16
    target_cell_col_map = 17
    target_cell_col_sql = 18
    sql = """
        SELECT OWNER, TABLE_NAME, T_COMMENTS, COLUMN_ID, COLUMN_NAME,
               C_COMMENTS, PK_POSITION, NULLABLE,
               DATA TYPE, DATA_LENGTH, DATA PRECISION, DATA SCALE, DATA_DEFAULT,
               PK_CONSTRAINT_NAME
         FROM HNF_TAB_COL_LIST A
        WHERE OWNER = 'RTMPKG' ORDER BY 1, 2, 3, 4
    """

    strVarName = ''
    list_col = []
    col_seq = 0

    cursor.execute(sql)
    rows = cursor.fetchall()
    for i, row in enumerate(rows):
        owner = row[0]
        table_id = row[1]
        table_name = row[2]

        col_seq = row[3]

        col_id = row[4]
        col_name = row[5]
        pk_yn = row[6]
        col_type = row[8]
        col_len = row[9]
        col_decimal_point = row[11]

        list_col_infos = [table_id, col_id, col_name, pk_yn, col_type.col_len, col_decimal_point]
        list_col.append(list_col_infos)

        # ------------------------------------------------------------------------------
        # 1. 변수명 생성하기
        # ------------------------------------------------------------------------------
        strVarName = self.f_make_var_name(col_id, col_name, col_type, col_decimal_point)

        # ------------------------------------------------------------------------------
        # 2. map에 넣는 구문 생성하기
        # ------------------------------------------------------------------------------
        strMapCode = self.f_make_var_map(col_id, col_name, col_type, col_decimal_point)

        # ------------------------------------------------------------------------------
        # 3. 변수명과 map 저장 구문을 db에 지장한다.
        # ------------------------------------------------------------------------------
        args = (table_id, table_name, col_seq, col_id, col_name, strVarName, strMapCode)
        self.f_save_var(cursor, args)

        # ------------------------------------------------------------------------------
        # Next Row의 테이블 id 확인 후 새로운 테이블 id를 만난 경우
        # ------------------------------------------------------------------------------
        if i+1 < len(rows): next_table_id = rows[i+1][1]
        if table_id != next_table_id or i+1 == len(rows): #zero base, max rowss
            # ------------------------------------------------------------------------------
            # * 2. insert sql 문 생성하기
            # ------------------------------------------------------------------------------
            myQuery = self.f_make_insert_sql(table_id, table_name, list_col)

            # oracle에 저장한다.
            self.f_save_sql_java_code(cursor, table_id, 'insert', myQuery)
            # 리스트 초기화.

            list_col = 0

    # oracle close
    cursor.execute("commit")
    cursor.close()
    conn.close()

    QMessageBox.Information(self, "알림", "작업이 종료되었습니다.", QMessageBox.Yes)




    #------------------------------------------------------------------------------
#* 테이블 칼럼으로 변수 생성
#------------------------------------------------------------------------------
def btnExcelJavaVar_clicked(self):

    fname = QFileDialog.getOpenFileName(
            self, '파일 선택', '',
            '모든 파일 (*.*);')[0]
    if not fname:
        return
    dsn = cx_Oracle.makedsn("localhost", 1521, "ord")
    conn = cx_Oracle.connect("han", "han", dsn)
    cursor = conn.cursor()

    # fname = r"C:03.프로젝트₩22.유렉스 포지션 XML 생성\Murex_Generator_Mapping 20210720.xlsx"
    wb = openpyxl.load_workbook(fname)
    ws_names = wb.sheetnames

    ws = wb['데이블정의서']

    #* 칼럼 길이
    col_id = col_name = pk_yn = col_type=''
    col_len = col_decimal_point = 0
    # 테이블 id
    table_id = ''
    table_name = ''


    target_type = ''
    target_cell_col_var = 16
    target_cell_col_map = 17
    target_cell_col_sql = 18


    strVarName = ''
    list_col = 0
    col_seq = 0
    for row in range(6, ws.max_row+1):
        col_seq += 1

        table_id = ws.cell(row, 5).value # 테이블 id
        table_name = ws.cell(row, 6).value # 테이블 name

        col_id = ws.cell(row, 7).value # 칼럼 id
        col_name = ws.cell(row, 8).value
        pk_yn = ws.cell(row, 9).value
        col_type = ws.cell(row, 13).value
        col_len = ws.cell(row, 14).value
        col_decimal_point = ws.cell(row, 15).value # 소수점 길이

        list_col_infos = [table_id, col_id,col_name,pk_yn,col_type.col_len,col_decimal_point]
        list_col.append(list_col_infos)


        #------------------------------------------------------------------------------
        # 1. 변수명 생성하기
        #------------------------------------------------------------------------------
        strVarName = self.f_make_var_name(col_id, col_name, col_type, col_decimal_point)
        ws.cell(row, target_cell_col_var).value = strVarName

        #------------------------------------------------------------------------------
        # 2. map에 넣는 구문 생성하기
        #------------------------------------------------------------------------------
        strMapCode = self.f_make_var_map(col_id, col_name, col_type, col_decimal_point)
        ws.cell(row, target_cell_col_map).value = strMapCode

        #------------------------------------------------------------------------------
        # 3. 변수명과 map 저장 구문을 db에 지장한다.
        #------------------------------------------------------------------------------
        args = (table_id, table_name, col_seq, col_id, col_name, strVarName, strMapCode)
        self.f_save_var(cursor, args)


        #------------------------------------------------------------------------------
        # Next Row의 테이블 id 확인 후 새로운 테이블 id를 만난 경우
        # ------------------------------------------------------------------------------
        next_table_id = ws.cell(row+1, 5).value # next 테이블 id
        if table_id != next_table_id or row == ws.max_row:
            #------------------------------------------------------------------------------
            # * 2. insert sql 문 생성하기
            #------------------------------------------------------------------------------
            myQuery = self.f_make_insert_sql(table_id, table_name, list_col)


            #------------------------------------------------------------------------------
            # * sql문을 cell에 넣는다.
            #------------------------------------------------------------------------------
            ws.cell(row, target_cell_col_sql).value = myQuery

            # oracle에 저장한다.
            self.f_save_insert_sql(cursor, table_id, myQuery)
            #리스트 초기화.

            list_col = 0
            col_seq = 0

    wb.save(fname)
    # oracle close
    cursor.execute("commit")
    cursor.close()
    conn.close()

    QMessageBox.Information(self, "알림","작업이 종료되었습니다.", QMessageBox.Yes)

def f_save_insert_sql(self, cursor, table_id, myQuery):
    # args = (table_id, table_name, col seq, cold, col_name, stivar Name, strMap Code)

    sql = "DELETE FROM JAVA_TBL_SQL_INSERT WHERE TABLE_NAME = :1"
    cursor.execute(sql, (table_id,))
    lines = myQuery.splitlines
    for seq, line in enumerate(lines):
        # insert
        sql = """
            VALUES(:1,:2,:3)
            INSERT INTO JAVA_TBL_SQL_INSERT (TABLE_NAME, seq, sql)
        """
        cursor.execute(sql, (table_id, seq, line))

#db 에 저장한다.
def f_save_var(self, cursor, args):
    #args = (table_id, table name, col_seq, col_id, col_name, strVarName, strMapCode)
    seq = args[2]

    if seq == 1:
        sql = "DELETE FROM JAVA_TBL_COL VAR WHERE TABLE_NAME = :1"
        cursor.execute(sql, (args[0], ))
        # insert
        sql = """
            INSERT INTO JAVA_TBL_COL VAR (TABLE_NAME, T_COMMENTS, SEQ
            COLUMN_NAME, C_COMMENTS, COL_VAR, COL_MAP)
            VALUES(:1, :2, :3, :4, :5, :6, :7)
        """
        cursor.execute(sql, args)

# ------------------------------------------------------------------------------
# map에 넣는 구문 생성하기
#------------------------------------------------------------------------------
def f_make_var_map(self, col_id, col_name, col_type, col_decimal_point):
    col_id = col_id.strip()
    strSource = " + col_id + "


    strCode = ''
    if col_id == "INS_USID" or col_id == "LST_CHG_USID":
        strLeft = ''
        strRight = ''
        strSource = "EOD_BATCH"

    elif col_id == "INS_DTM" or col_id == "LST_CHG_DTM":

        strLeft = ''
        strRight = ''
        strSource = 'KFDSCom.getCurrentDate+""+ KFDSCom.getCurrentTime()'

    elif col_type.find('CHAR') >= 0:

        strLeft = '(String) map.get('
        strRight = ')'

    elif col_type.find('DATE') >= 0:
        strLeft = 'dateNumber((String) map.get('
        strRight = ')'

    elif col_type.find('NUMBER') >= 0:
        if col_decimal_point: # double

            strLeft = 'Double.parseDouble((String)map.get('
            strRight = '))'
        else: #long

            strleft = 'Long parselong((String)map.get('
            strRight = '))'

    else:
        strLeft = '(string)map.get("'
        strRight = ')'

    # public String RECV_DATE ; // NOT NULL


    strMapCode = f'entity.{col_id: <30} = {strLeft}{strSource}{strRight}'
    strMapCode = f'{strMapCode:<100} // {col_name}'
    return strMapCode

#------------------------------------------------------------------------------
# *java insert 문 생성하기
#------------------------------------------------------------------------------

def _make_insert_sal(self, table_id, table_name, list_col):
    col_count = len(list_col)
    if col_count == 0:
        return None

    mySpace = ''
    #------------------------------------------------------------------------------
    #* 1. insert 문
    #------------------------------------------------------------------------------
    str1 = ''
    str1 = str1 + f'qry.append("INSERT INTO HRDB. {table_id} "); // {table_name}\n'
    str1 = str1 + f'qry.append("({mySpace:<30)"};Wn'

    # ------------------------------------------------------------------------------
    #* 1.1 리스트 순회
    # ------------------------------------------------------------------------------
    comma = ''
    for i, col in enumerate(list_col):
        col_id = col[1]
        col_name = col[2]
        if i <col_count - 1: # zero base
            comma = ','
        else:
            comma =' ' #콤마 제거 버전

        #------------------------------------------------------------------------------
        # * 1.2 dp 칼럼명 나열
        #------------------------------------------------------------------------------
        str1 = str1 + f'qry.append("{col_id:<30}{comma}");' #콤마 제거버전
        str1 = str1 + f'\t //(col_name: <30)Wn'

    str1 = str1 + f'qry.append("){mySpace:< 30}");\n'
    # ------------------------------------------------------------------------------
    # * 2. values 구문
    #------------------------------------------------------------------------------
    str2 = '\n\n'
    str2 = str2 + f'qry.append("VALUES ");\n'
    str2 = str2 + f'qry.append("({mySpace:<30});\n'

    dateQ = " to_date(?, 'yyyymmdd hh24:mi:ss')"
    normQ = " ?"


    #------------------------------------------------------------------------------
    # 2.1 리스트 순회
    #------------------------------------------------------------------------------
    for i, col in enumerate(list_col):
        col_id = col[1]
        col_name = col[2]
        col_type = col[4]
        Q = normQ
        if col_type == 'DATE': #날짜는 to_date(?, 'yyyymmdd')와 같이 표현한다.
            Q = dateQ
        if i < col_count - 1: #zero base
            comma = ','
        else:
            comma = ' ' # 콤마 제거 비전

        # ------------------------------------------------------------------------------
        # * 2.2 values 나열
        # ------------------------------------------------------------------------------
        str2 = str2 + f'qry.append("{Q:<30}{comma}");'
        str2 = str2 + f'\t// {col_id:<30}\n'

    str2 = str2 + f'qry.append("){mySpace: < 30}");\n'
    sql = str1 + str2
    return sql


#------------------------------------------------------------------------------
#* 변수명 생성하기
#------------------------------------------------------------------------------
def f_make_var_name(self, col_id, col_name, col_type, col_decimal_point):
    if col_type.find('CHAR') >= 0:
        target_type = 'String'
    elif col_type.find('DATE') >= 0:
        target_type = 'String'
    elif col_type.find('NUMBER') >= 0:
        if col_decimal_point:
            target_type = 'double'
        else:
            target_type = 'long'
    else:
        target_type = 'String'

    # public String RECV_DATE; // NOT NULL
    strVarName = f"public {target_type:<9}{col_id:<30};\t// {col_name}"
    return strVarName